import csv
from pathlib import Path
from typing import Protocol

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

from brightsidebudget.account.account import Account
from brightsidebudget.account.account_type import AccountType

HEADER : list[str] = ["Compte", "Type", "Groupe", "Sous-groupe", "Numéro"]

class IAccountRepository(Protocol):
    """Protocol for account repository operations."""

    def write_accounts(self, *, accounts: list[Account], destination: Path):
        """Write accounts."""
        ...

    def get_accounts(self, source: Path) -> list[Account]:
        """Retrieve accounts."""
        ...

class CsvAccountRepository(IAccountRepository):
    """Repository for managing accounts in CSV format."""

    def write_accounts(self, *,
                       accounts: list['Account'],
                       destination: Path):
        accounts = sorted(accounts, key=lambda a: a.sort_key())
        with open(destination, "w", encoding="utf-8") as file:
            writer = csv.DictWriter(file, 
                                    fieldnames=HEADER,
                                    lineterminator="\n")
            writer.writeheader()
            for a in accounts:
                writer.writerow(a.to_dict())

    def get_accounts(self, source: Path) -> list['Account']:
        ls = []
        with open(source, "r", encoding="utf-8", newline='') as file:
            for row in csv.DictReader(file):
                a = Account.from_dict(row)
                ls.append(a)
        return ls

class ExcelAccountRepository(IAccountRepository):
    """Repository for managing accounts in Excel format."""
    
    def write_accounts(self, *,
                       accounts: list['Account'],
                       destination: Path):
        accounts = sorted(accounts, key=lambda a: a.sort_key())
        # Load existing workbook or create new one
        wb = self._get_or_create_workboot(destination)
        ws = wb["Comptes"]
        
        # Add data
        ws.append(HEADER)
        for a in accounts:
            ws.append([
                a.name,
                a.type.name,
                a.group,
                a.subgroup,
                str(a.number)
            ])
        
        # Update the table range
        last_row = ws.max_row
        last_col = len(HEADER)
        new_range = f"A1:{get_column_letter(last_col)}{last_row}"
        self._set_account_table(ws, new_range)
        
        wb.save(destination)

    def _get_or_create_workboot(self, destination: Path) -> Workbook:
        """
        Get or create a workbook with a "Comptes" worksheet.
        """
        if destination.exists():
            wb = openpyxl.load_workbook(destination)
            if "Comptes" in wb.sheetnames:
                ws = wb["Comptes"]
                # Clear all existing data
                ws.delete_rows(1, ws.max_row)
            else: 
                # Create new "Comptes" worksheet
                ws = wb.create_sheet("Comptes")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Failed to create a new worksheet.")
            ws.title = "Comptes"
            ws.sheet_view.showGridLines = False
        return wb

    def _set_account_table(self, ws: Worksheet, range: str):
        if ws.tables:
            # Get the first table (assuming only one table in this sheet)
            existing_table = list(ws.tables.values())[0]
            existing_table.ref = range
        else:
            # Create new table with name "Comptes"
            table = Table(displayName="Comptes", ref=range)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
    
    def get_accounts(self, source: Path) -> list['Account']:
        wb = openpyxl.load_workbook(source, data_only=True)
        if "Comptes" in wb.sheetnames:
            ws = wb["Comptes"]
        else:
            raise ValueError("Worksheet 'Comptes' not found in the workbook.")
        accounts: list[Account] = []
        # assume header is in row 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                raise ValueError("Row does not contain enough columns for Account data.")
            name, type_, group, subgroup, number = row
            acc = Account(
                name=str(name),
                type=AccountType(name=str(type_)), # type: ignore
                group=str(group) if group else "",
                subgroup=str(subgroup) if subgroup else "",
                number=int(str(number))
            )
            accounts.append(acc)
        return accounts