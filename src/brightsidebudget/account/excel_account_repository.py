from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from brightsidebudget.account.account import Account
from brightsidebudget.account.account_type import AccountType
from brightsidebudget.utils.excel_utils import get_or_create_clean_ws, load_or_create_workbook, set_table_range

HEADER : list[str] = ["Compte", "Type", "Groupe", "Sous-groupe", "Numéro"]

class ExcelAccountRepository():
    """Repository for managing accounts in Excel format."""
    
    def write_accounts(self, *,
                       accounts: list['Account'],
                       destination: Path):
        wb = load_or_create_workbook(destination)
        ws = get_or_create_clean_ws(wb, "Comptes")
        self.write_accounts_worksheet(accounts=accounts, ws=ws)
        wb.save(destination)

    def write_accounts_worksheet(self, *,
                       accounts: list['Account'],
                       ws: Worksheet):
        accounts = sorted(accounts, key=lambda a: a.sort_key())
        
        # Add data
        ws.append(HEADER)
        for a in accounts:
            ws.append([
                a.name,
                a.type.name,
                a.group,
                a.subgroup,
                a.number
            ])
        
        # Update the table range
        last_row = ws.max_row
        last_col = len(HEADER)
        new_range = f"A1:{get_column_letter(last_col)}{last_row}"
        set_table_range(ws, "Comptes", new_range)

    def get_accounts(self, source: Path) -> list['Account']:
        wb = openpyxl.load_workbook(source, data_only=True)
        if "Comptes" in wb.sheetnames:
            return self.get_accounts_worksheet(wb["Comptes"])
        else:
            raise ValueError("Worksheet 'Comptes' not found in the workbook.")

    def get_accounts_worksheet(self, ws: Worksheet) -> list['Account']:
        accounts: list[Account] = []
        # assume header is in row 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 5:
                raise ValueError("Row does not contain enough columns for Account data.")
            name, type_, group, subgroup, number = row
            acc = Account(
                name=str(name),
                type=AccountType(name=str(type_)), # type: ignore
                group=str(group) if group else "",
                subgroup=str(subgroup) if subgroup else "",
                number=int(str(number))
            )
            accounts.append(acc)
        return accounts