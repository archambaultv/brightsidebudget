from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

def load_or_create_workbook(destination: Path) -> Workbook:
    """
    Load an existing workbook or create a new one if it doesn't exist.
    """
    if destination.exists():
        return openpyxl.load_workbook(destination)
    else:
        return openpyxl.Workbook()

def get_or_create_clean_ws(wb: Workbook, ws_name: str) -> Worksheet:
    """
    Get or create a workbook with a clean (no data) "ws_name" worksheet.
    """
    if ws_name in wb.sheetnames:
        ws = wb[ws_name]
        # Clear all existing data
        ws.delete_rows(1, ws.max_row)
    else: 
        # Create new ws_name worksheet
        ws = wb.create_sheet(ws_name)
    return ws

def set_table_range(ws: Worksheet, range_name: str, range: str):
    if ws.tables and range_name in ws.tables:
        existing_table = ws.tables[range_name]
        existing_table.ref = range
    else:
        # Create new table with name "range_name"
        table = Table(displayName=range_name, ref=range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                showLastColumn=False, showRowStripes=True,
                                showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)