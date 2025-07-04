from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from brightsidebudget.account.account import Account
from brightsidebudget.bassertion.bassertion import BAssertion
from brightsidebudget.utils.excel_utils import get_or_create_clean_ws, load_or_create_workbook, set_table_range

HEADER = ["Date", "Compte", "Solde", "Commentaire"]

class ExcelBAssertionRepository():
    """Repository for managing BAssertions in Excel format."""

    def write_bassertions(self, *, bassertions: list[BAssertion], destination: Path):
        wb = load_or_create_workbook(destination)
        ws = get_or_create_clean_ws(wb, "Soldes")
        self.write_bassertions_worksheet(bassertions=bassertions, ws=ws)
        wb.save(destination)

    def write_bassertions_worksheet(self, *, bassertions: list[BAssertion], ws: Worksheet):
        bs = sorted(bassertions, key=lambda b: b.sort_key())

        ws.append(HEADER)
        for b in bs:
            ws.append([
                b.date,
                b.account.name,
                float(b.balance),
                b.comment
            ])

        last_row = ws.max_row
        last_col = len(HEADER)
        new_range = f"A1:{get_column_letter(last_col)}{last_row}"
        set_table_range(ws, "Soldes", new_range)

    def get_bassertions(self, source: Path, accounts: dict[str, Account]) -> list[BAssertion]:
        wb = openpyxl.load_workbook(source, data_only=True)
        if "Soldes" in wb.sheetnames:
            return self.get_bassertions_worksheet(wb["Soldes"], accounts)
        else:
            raise ValueError("Worksheet 'Soldes' not found in the workbook.")
        
    def get_bassertions_worksheet(self, ws: Worksheet, accounts: dict[str, Account]) -> list[BAssertion]:
        bassertions: list[BAssertion] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                raise ValueError("Row does not contain enough columns for BAssertion data.")
            date_str, compte, solde, commentaire = row
            bassertion = BAssertion(
                date=date_str, # type: ignore
                account=accounts[str(compte)],
                balance=solde, # type: ignore
                comment=commentaire if commentaire else "" # type: ignore
            )
            bassertions.append(bassertion)
        return bassertions