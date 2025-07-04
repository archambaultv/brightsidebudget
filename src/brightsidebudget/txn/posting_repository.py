from pathlib import Path
from typing import Protocol

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook

from brightsidebudget.account.account import Account
from brightsidebudget.txn.posting import Posting

HEADER = ["No txn", "Date", "Compte", "Montant", "Date du relevé", "Commentaire",
          "Description du relevé"]

class IPostingRepository(Protocol):
    """Protocol for posting repository operations."""

    def write_postings(self, *, postings: list[Posting], destination: Path):
        """Write postings."""
        ...

    def get_postings(self, source: Path, accounts: dict[str, Account]) -> list[Posting]:
        """Retrieve postings."""
        ...

class ExcelPostingRepository(IPostingRepository):
    """Repository for managing postings in Excel format."""
    
    def write_postings(self, *, 
                       postings: list[Posting], 
                       destination: Path):
        """Write postings to Excel file."""
        ps = sorted(postings, key=lambda p: p.sort_key())
        
        # Load existing workbook or create new one
        wb = self._get_or_create_workbook(destination)
        ws = wb["Txns"]
        
        # Add data
        ws.append(HEADER)
        for p in ps:
            ws.append([
                str(p.txn_id),
                str(p.date),
                p.account.name,
                str(p.amount),
                str(p.stmt_date),
                p.comment,
                p.stmt_desc
            ])
        
        # Update the table range
        last_row = ws.max_row
        last_col = len(HEADER)
        new_range = f"A1:{get_column_letter(last_col)}{last_row}"
        self._set_posting_table(ws, new_range)
        
        wb.save(destination)

    def _get_or_create_workbook(self, destination: Path) -> Workbook:
        """
        Get or create a workbook with a "Postings" worksheet.
        """
        if destination.exists():
            wb = openpyxl.load_workbook(destination)
            if "Txns" in wb.sheetnames:
                ws = wb["Txns"]
                # Clear all existing data
                ws.delete_rows(1, ws.max_row)
            else: 
                # Create new "Txns" worksheet
                ws = wb.create_sheet("Txns")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Failed to create a new worksheet.")
            ws.title = "Txns"
            # Set gridlines view to false for new workbook
            ws.sheet_view.showGridLines = False
        return wb

    def _set_posting_table(self, ws: Worksheet, range: str):
        if ws.tables:
            # Get the first table (assuming only one table in this sheet)
            existing_table = list(ws.tables.values())[0]
            existing_table.ref = range
        else:
            # Create new table with name "Txns"
            table = Table(displayName="Txns", ref=range)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)
    
    def get_postings(self, source: Path, accounts: dict[str, Account]) -> list[Posting]:
        """Retrieve postings from Excel file."""
        wb = openpyxl.load_workbook(source, data_only=True)
        if "Txns" in wb.sheetnames:
            ws = wb["Txns"]
        else:
            raise ValueError("Worksheet 'Txns' not found in the workbook.")
        postings: list[Posting] = []
        # assume header is in row 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                raise ValueError("Row does not contain enough columns for Posting data.")
            txn_id, date_str, compte, montant, stmt_date_str, commentaire, description = row
            
            try:
                # Convert to dict format expected by Posting.from_dict
                posting = Posting(
                    txn_id=txn_id, # type: ignore
                    date=date_str,  # type: ignore
                    account=accounts[str(compte)],
                    amount=montant,  # type: ignore
                    stmt_date=stmt_date_str, # type: ignore
                    comment=commentaire if commentaire else "",  # type: ignore
                    stmt_desc=description if description else ""  # type: ignore
                )
            except Exception as e:
                raise ValueError(f"Error processing row {row}") from e
            postings.append(posting)
        return postings
