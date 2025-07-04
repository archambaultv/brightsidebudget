from pathlib import Path
from typing import Protocol
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from brightsidebudget.account.account import Account
from brightsidebudget.bassertion.bassertion import BAssertion

HEADER = ["Date", "Compte", "Solde", "Commentaire"]

class IBAssertionRepository(Protocol):
    """Protocol for BAssertion repository operations."""

    def write_bassertions(self, *, bassertions: list[BAssertion], destination: Path):
        ...

    def get_bassertions(self, source: Path, accounts: dict[str, Account]) -> list[BAssertion]:
        ...


class ExcelBAssertionRepository(IBAssertionRepository):
    """Repository for managing BAssertions in Excel format."""

    def write_bassertions(self, *, bassertions: list[BAssertion], destination: Path):
        bs = sorted(bassertions, key=lambda b: b.sort_key())
        wb = self._get_or_create_workbook(destination)
        ws = wb["Soldes"]

        ws.append(HEADER)
        for b in bs:
            ws.append([
                str(b.date),
                b.account.name,
                str(b.balance),
                b.comment
            ])

        last_row = ws.max_row
        last_col = len(HEADER)
        new_range = f"A1:{get_column_letter(last_col)}{last_row}"
        self._set_bassertion_table(ws, new_range)

        wb.save(destination)

    def _get_or_create_workbook(self, destination: Path) -> Workbook:
        if destination.exists():
            wb = openpyxl.load_workbook(destination)
            if "Soldes" in wb.sheetnames:
                ws = wb["Soldes"]
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet("Soldes")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            if ws is None:
                raise ValueError("Failed to create a new worksheet.")
            ws.title = "Soldes"
            ws.sheet_view.showGridLines = False
        return wb

    def _set_bassertion_table(self, ws: Worksheet, range: str):
        if ws.tables:
            existing_table = list(ws.tables.values())[0]
            existing_table.ref = range
        else:
            table = Table(displayName="Soldes", ref=range)
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True,
                                   showColumnStripes=False)
            table.tableStyleInfo = style
            ws.add_table(table)

    def get_bassertions(self, source: Path, accounts: dict[str, Account]) -> list[BAssertion]:
        wb = openpyxl.load_workbook(source, data_only=True)
        if "Soldes" in wb.sheetnames:
            ws = wb["Soldes"]
        else:
            raise ValueError("Worksheet 'Soldes' not found in the workbook.")
        bassertions: list[BAssertion] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 4:
                raise ValueError("Row does not contain enough columns for BAssertion data.")
            date_str, compte, solde, commentaire = row
            bassertion = BAssertion(
                date=date_str, # type: ignore
                account=accounts[str(compte)],
                balance=solde, # type: ignore
                comment=commentaire if commentaire else "" # type: ignore
            )
            bassertions.append(bassertion)
        return bassertions